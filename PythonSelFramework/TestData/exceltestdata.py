import openpyxl

excel = openpyxl.load_workbook("C:\\Users\\HP\\PycharmProjects\\PythonLearning\\Python_Testdata.xlsx")
sheet = excel.active
data = sheet.cell(row=1, column =2)
print(data.value)
print("*********************************")
sheet.cell(row=6, column=2).value = "Revathi"
print(sheet.cell(row=6, column=2).value)
Dict = {}

print(sheet.cell(row=1, column = 3).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet['A5'].value)


print("*********************************")

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column =1).value == "TC 4":
        for j in range(1, sheet.max_column+1):
          print(sheet.cell(row=i, column = j).value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column =1).value == "TC 3":
        for j in range(1, sheet.max_column+1):
            Dict[sheet.cell(row=1, column = j).value] = sheet.cell(row=i, column =j).value

print(Dict)
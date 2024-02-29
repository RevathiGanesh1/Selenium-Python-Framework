
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl



def excel_test_data(filePath, search, colname, new_Value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}


#iterating the column which has heading "price"
    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == colname:
            Dict["col"] = i

#iterating row which has value "Apple" and updating its price
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == search:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_Value
    book.save(file_path)


fruit_name = "Apple"
newValue = 555
file_path = "C:\\Users\\HP\\Downloads\\download.xlsx"
#service_obj = Service()
#driver = webdriver.Chrome(service = service_obj)
driver = webdriver.Chrome()
driver.implicitly_wait(5)

#downloading excel in webpage
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.find_element(By.CSS_SELECTOR,".button").click()

#editing excel file column
excel_test_data(file_path, fruit_name, "price", newValue)

#uploading to webpage
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
columnAttribute = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+columnAttribute+"-undefined']").text
print(actual_price)

assert actual_price == newValue



